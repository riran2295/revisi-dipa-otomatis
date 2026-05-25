import streamlit as st
import pdfplumber
import re
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from wilayah import tentukan_wilayah

# ==========================================
# KONFIGURASI DNS & HALAMAN
# ==========================================
st.set_page_config(page_title="revisi dipa otomatis", page_icon="🚀", layout="centered")

# --- FUNGSI BACA GAMBAR JADI BACKGROUND ---
def get_base64_of_bin_file(bin_file):
    try:
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except Exception:
        return None

# Ubah bg.png jadi kode base64
bg_base64 = get_base64_of_bin_file("bg.png")

if bg_base64:
    bg_css = f"""
    <style>
    .stApp {{
        background-image: url("data:image/png;base64,{bg_base64}");
        background-repeat: repeat;
        background-attachment: fixed;
    }}
    header[data-testid="stHeader"] {{
        background-color: transparent !important;
    }}
    </style>
    """
else:
    bg_css = """
    <style>
    .stApp {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        background-attachment: fixed;
        background-size: cover;
    }
    header[data-testid="stHeader"] {
        background-color: transparent !important;
    }
    </style>
    """

# Injeksi CSS Gabungan (UI Premium & Glassmorphism)
st.markdown(bg_css + """
    <style>
    .block-container {
        background-color: rgba(255, 255, 255, 0.90);
        padding: 3rem 2rem !important;
        border-radius: 20px;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.15);
        margin-top: 2rem;
        margin-bottom: 5rem;
    }
    .title-text {
        text-align: center;
        font-weight: 800;
        font-size: 2.2rem;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
        color: #0f172a;
    }
    .subtitle-text {
        text-align: center;
        color: #475569;
        margin-bottom: 1rem;
        font-weight: 500;
    }
    .warning-text {
        text-align: center;
        color: #e11d48; /* Warna merah tegas elegan */
        font-weight: 600;
        font-size: 1rem;
        margin-bottom: 2rem;
        background-color: #fff1f2;
        padding: 10px;
        border-radius: 8px;
        border: 1px solid #fecdd3;
    }
    div[data-testid="stFileUploadDropzone"] {
        background-color: #f8fafc;
        border: 2px dashed #94a3b8;
        border-radius: 12px;
        padding: 20px;
        transition: all 0.3s ease;
    }
    div[data-testid="stFileUploadDropzone"]:hover {
        border-color: #3b82f6;
        background-color: #eff6ff;
    }
    .footer {
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: rgba(15, 23, 42, 0.85);
        backdrop-filter: blur(5px);
        color: #f8fafc;
        text-align: center;
        padding: 12px;
        font-size: 13px;
        z-index: 100;
        letter-spacing: 1px;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# TAMPILAN HEADER (LOGO & JUDUL)
# ==========================================
col_logo1, col_logo2, col_logo3 = st.columns([1, 3, 1])
with col_logo2:
    try:
        st.image("logo.png", use_container_width=True)
    except Exception as e:
        st.warning("File 'logo.png' belum ada di folder. Silakan tambahkan dulu gambarnya.")

st.markdown('<div class="title-text">Automasi Reviu Revisi DIPA</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle-text">Upload puluhan file PDF Matriks Perubahan sekaligus, sistem akan otomatis mengekstrak selisih KRO dan meng-generate Kertas Kerja Excel.</div>', unsafe_allow_html=True)
# PENAMBAHAN TULISAN PERINGATAN FORMAT PDF
st.markdown('<div class="warning-text">⚠️ Hanya untuk file PDF matriks perubahan format 1 yang bukan hasil scan</div>', unsafe_allow_html=True)

# ==========================================
# LOGIKA UTAMA APLIKASI
# ==========================================
col_up1, col_up2, col_up3 = st.columns([1, 8, 1])
with col_up2:
    uploaded_files = st.file_uploader("Upload File PDF Matriks (Bisa pilih banyak file)", type=["pdf"], accept_multiple_files=True)

col_btn1, col_btn2, col_btn3 = st.columns([1, 3, 1])
with col_btn2:
    proses_btn = st.button("Proses & Generate Excel", type="primary", use_container_width=True)

if proses_btn:
    if not uploaded_files:
        st.warning("Pilih minimal 1 file PDF dulu ya!")
    else:
        with st.spinner(f'Lagi nge-ekstrak {len(uploaded_files)} file PDF...'):
            data_revisi = []
            pola_angka = r'-?\d+(?:,\d{3})*(?:\.\d+)?'

            for file_pdf in uploaded_files:
                try:
                    with pdfplumber.open(file_pdf) as pdf:
                        satker_aktif = "Belum dapet Satker"
                        jenis_satker = "Daerah" 
                        unit_org = ""  
                        
                        for halaman in pdf.pages:
                            teks = halaman.extract_text()
                            if not teks: continue
                                
                            baris_teks = teks.split('\n')
                            kro_aktif = None 
                            hitung_baris = 0
                            baris_uang = ""
                            
                            for baris in baris_teks:
                                baris_upper = baris.upper()
                                
                                if "UNIT ORGANISASI" in baris_upper:
                                    pecah = re.split(r'(?i)UNIT\s*ORGANISASI', baris)
                                    nama_clean = re.sub(r'^[\s\:\-]*', '', pecah[-1])
                                    nama_clean = re.sub(r'^\(\w+\)\s*', '', nama_clean)
                                    unit_org = re.sub(r'^\d+[\s\.\-]*', '', nama_clean).strip()
                                    
                                    if satker_aktif == "Belum dapet Satker" or "KEMENTERIAN" in satker_aktif.upper():
                                        satker_aktif = unit_org
                                        jenis_satker = "Pusat"
                                    continue

                                elif "SATUAN KERJA" in baris_upper or "SATKER" in baris_upper:
                                    pecah = re.split(r'(?i)SATUAN\s*KERJA|SATKER', baris)
                                    nama_clean = re.sub(r'^[\s\:\-]*', '', pecah[-1])
                                    nama_clean = re.sub(r'^\(\w+\)\s*', '', nama_clean)
                                    nama_clean = re.sub(r'^\d+[\s\.\-]*', '', nama_clean).strip()
                                    
                                    if nama_clean:
                                        if ("KEMENTERIAN" in nama_clean.upper() or "KEMEN" in nama_clean.upper()) and unit_org != "":
                                            satker_aktif = unit_org
                                            jenis_satker = "Pusat"
                                        else:
                                            satker_aktif = nama_clean
                                            if any(kw in satker_aktif.upper() for kw in ["KANTOR PERTANAHAN", "KANTAH", "KANTOR WILAYAH", "KANWIL"]):
                                                jenis_satker = "Daerah"
                                            else:
                                                jenis_satker = "Pusat"
                                    continue

                                elif any(kw in baris_upper for kw in ["DIREKTORAT", "DITJEN", "INSPEKTORAT", "SEKRETARIAT", "BADAN ", "BPSDM", "BIRO ", "STPN", "PUSAT"]):
                                    if "KEMENTERIAN" not in baris_upper and (satker_aktif == "Belum dapet Satker" or "KEMENTERIAN" in satker_aktif.upper()):
                                        satker_aktif = re.sub(r'^\d+[\s\.\-]*', '', baris).strip()
                                        jenis_satker = "Pusat" 
                                    continue

                                elif any(kw in baris_upper for kw in ["KANTOR PERTANAHAN", "KANTAH", "KANTOR WILAYAH", "KANWIL"]):
                                    if satker_aktif == "Belum dapet Satker" or "KEMENTERIAN" in satker_aktif.upper() or jenis_satker == "Pusat":
                                        satker_aktif = re.sub(r'^\d+[\s\.\-]*', '', baris).strip()
                                        jenis_satker = "Daerah" 
                                    continue

                                kro_match = re.search(r'^(\d{4}\.[A-Z]{3})', baris)
                                if kro_match:
                                    kro_aktif = kro_match.group(1)
                                    hitung_baris = 0
                                    continue 
                                
                                if kro_aktif:
                                    hitung_baris += 1
                                    if hitung_baris == 1:
                                        baris_uang = baris
                                    elif hitung_baris == 2:
                                        baris_vol = baris
                                        angka_uang = re.findall(pola_angka, baris_uang)
                                        
                                        baris_vol_clean = baris_vol.replace("U M B E R", "UMBER").replace("S U M", "SUM")
                                        angka_vol = re.findall(pola_angka, baris_vol_clean)
                                        
                                        if len(angka_uang) >= 6 and len(angka_vol) >= 3:
                                            selisih_jml = int(angka_uang[4].replace(',', ''))
                                            selisih_blk = int(angka_uang[5].replace(',', ''))
                                            selisih_vol = float(angka_vol[2].replace(',', ''))
                                            
                                            if selisih_jml != 0 or selisih_blk != 0 or selisih_vol != 0:
                                                data_revisi.append({
                                                    'satker': satker_aktif,
                                                    'jenis': jenis_satker, 
                                                    'KRO': kro_aktif,
                                                    'semula_vol': float(angka_vol[0].replace(',', '')),
                                                    'semula_jml': int(angka_uang[0].replace(',', '')),
                                                    'semula_blk': int(angka_uang[1].replace(',', '')),
                                                    'menjadi_vol': float(angka_vol[1].replace(',', '')),
                                                    'menjadi_jml': int(angka_uang[2].replace(',', '')),
                                                    'menjadi_blk': int(angka_uang[3].replace(',', ''))
                                                })
                                        kro_aktif = None
                except Exception as e:
                    st.error(f"Gagal membaca file {file_pdf.name}: {e}")

            if len(data_revisi) > 0:
                wb = Workbook()
                sheet = wb.active
                sheet.title = "Reviu Revisi DIPA"
                sheet.views.sheetView[0].showGridLines = True

                headers = [
                    "No", "Wilayah", "Nama Satuan Kerja", "KRO", 
                    "Semula Vol", "Semula Jumlah", "Semula Blokir", 
                    "Menjadi Vol", "Menjadi Jumlah", "Menjadi Blokir",
                    "Selisih Vol", "Selisih Jumlah", "Selisih Blokir"
                ]

                font_header = Font(name="Arial", size=10, bold=True)
                kuning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_align = Alignment(horizontal="left", vertical="center")
                right_align = Alignment(horizontal="right", vertical="center")
                border_tipis = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )

                sheet.row_dimensions[4].height = 28
                for col_idx, header_text in enumerate(headers, 1):
                    cell = sheet.cell(row=4, column=col_idx)
                    cell.value = header_text
                    cell.font = font_header
                    cell.fill = kuning_fill
                    cell.alignment = center_align
                    cell.border = border_tipis

                baris_mulai = 5
                for indeks, d in enumerate(data_revisi):
                    baris = baris_mulai + indeks
                    sheet.row_dimensions[baris].height = 20
                    
                    satker_rapi = d['satker'].title().replace("Kab.", "Kabupaten")
                    wilayah = tentukan_wilayah(d['satker'], d['jenis']) 
                    
                    sheet.cell(row=baris, column=1, value=indeks + 1).alignment = center_align
                    sheet.cell(row=baris, column=2, value=wilayah).alignment = left_align
                    sheet.cell(row=baris, column=3, value=satker_rapi).alignment = left_align
                    sheet.cell(row=baris, column=4, value=d['KRO']).alignment = center_align
                    
                    sheet.cell(row=baris, column=5, value=d['semula_vol'])
                    sheet.cell(row=baris, column=6, value=d['semula_jml'])
                    sheet.cell(row=baris, column=7, value=d['semula_blk'])
                    sheet.cell(row=baris, column=8, value=d['menjadi_vol'])
                    sheet.cell(row=baris, column=9, value=d['menjadi_jml'])
                    sheet.cell(row=baris, column=10, value=d['menjadi_blk'])

                    sheet.cell(row=baris, column=11, value=f"=H{baris}-E{baris}")
                    sheet.cell(row=baris, column=12, value=f"=I{baris}-F{baris}")
                    sheet.cell(row=baris, column=13, value=f"=J{baris}-G{baris}")
                    
                    for col_idx in range(1, 14):
                        c = sheet.cell(row=baris, column=col_idx)
                        c.border = border_tipis
                        c.font = Font(name="Arial", size=10)
                        
                        if col_idx >= 5: 
                            c.number_format = '#,##0'
                            c.alignment = right_align

                for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = re.sub(r'[^A-Z]', '', col[0].coordinate)
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

                output_excel = io.BytesIO()
                wb.save(output_excel)
                output_excel.seek(0)
                
                st.balloons()
                st.success(f"Sukses! Menemukan total {len(data_revisi)} baris selisih.")
                
                col_dl1, col_dl2, col_dl3 = st.columns([1, 3, 1])
                with col_dl2:
                    st.download_button(
                        label="⬇️ Download Kertas Kerja Excel",
                        data=output_excel,
                        file_name="kertas_kerja_OTOMATIS.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
            else:
                st.info("Semua PDF sudah dicek, tapi tidak ditemukan selisih KRO.")

# Menampilkan Footer
st.markdown('<div class="footer">Developed by riranism</div>', unsafe_allow_html=True)
